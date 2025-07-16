from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl import Workbook
import os
import datetime
import re
import random
import json
from uuid import uuid4  # For unique IDs

app = Flask(__name__)

# Initialize data files
def init_data_files():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    # Product data
    product_filepath = 'data/artisandata.xlsx'
    if not os.path.exists(product_filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Product Name', 'Category', 'Description_EN', 'Description_TA', 'Quantity', 'Price', 'Timestamp'])
        wb.save(product_filepath)
    
    # Materials data
    materials_filepath = 'data/materials.json'
    if not os.path.exists(materials_filepath):
        with open(materials_filepath, 'w') as f:
            json.dump([], f)
    
    # Study resources
    study_filepath = 'data/study.json'
    if not os.path.exists(study_filepath):
        with open(study_filepath, 'w') as f:
            json.dump([], f)
    
    # Settings
    settings_filepath = 'data/settings.json'
    if not os.path.exists(settings_filepath):
        with open(settings_filepath, 'w') as f:
            json.dump({'language': 'EN', 'theme': 'light'}, f)
    
    return {
        'products': product_filepath,
        'materials': materials_filepath,
        'study': study_filepath,
        'settings': settings_filepath
    }

# AI Functions (Stubs with Tamil support)
def generate_description(product_name, lang):
    if lang == 'TA':
        return f"உயர்தர {product_name}, பாரம்பரிய கைவினைத் திறனால் உருவாக்கப்பட்டது"
    return f"Handcrafted {product_name} made with traditional techniques"

def predict_category(product_name):
    # Tamil product recognition
    tamil_keywords = {
        'புடவை': 'Textiles',
        'சேலை': 'Textiles',
        'மட்பாண்டம்': 'Pottery',
        'விளக்கு': 'Pottery',
        'செம்பு': 'Metalwork',
        'வெண்கலம்': 'Metalwork',
        'நகை': 'Jewelry'
    }
    
    # Check for Tamil keywords
    for keyword, category in tamil_keywords.items():
        if keyword in product_name:
            return category
    
    # English fallback
    if 'silk' in product_name.lower() or 'saree' in product_name.lower():
        return 'Textiles'
    elif 'pot' in product_name.lower() or 'vase' in product_name.lower():
        return 'Pottery'
    elif 'metal' in product_name.lower() or 'brass' in product_name.lower():
        return 'Metalwork'
    elif 'jewelry' in product_name.lower() or 'necklace' in product_name.lower():
        return 'Jewelry'
    return 'Handicrafts'

def extract_price(input_text):
    # Extract numbers with currency context
    matches = re.findall(r'(\d+)\s*(ரூபாய்|ரூ|rupees?|rs|₹)', input_text, re.IGNORECASE)
    if matches:
        return int(matches[-1][0])  # Take last match
    numbers = re.findall(r'\d+', input_text)
    return int(numbers[-1]) if numbers else 1000  # Default price

def extract_quantity(input_text):
    numbers = re.findall(r'\d+', input_text)
    return int(numbers[0]) if numbers else 1  # Default quantity

# Existing product routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add_product', methods=['POST'])
def add_product():
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        data = request.json
        product_name = data['name'].strip()
        quantity = int(data['quantity'])
        price = int(data['price'])
        
        # Validate inputs
        if not product_name or quantity <= 0 or price <= 0:
            return jsonify(success=False, error="Invalid input values"), 400
        
        category = predict_category(product_name)
        desc_en = generate_description(product_name, 'EN')
        desc_ta = generate_description(product_name, 'TA')
        
        # Generate unique ID
        new_id = str(uuid4())
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            new_id, 
            product_name, 
            category, 
            desc_en, 
            desc_ta, 
            quantity, 
            price, 
            timestamp
        ])
        
        wb.save(files['products'])
        return jsonify(success=True, id=new_id)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route('/get_products')
def get_products():
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                products.append({
                    'id': row[0],
                    'name': row[1],
                    'category': row[2],
                    'desc_en': row[3],
                    'desc_ta': row[4],
                    'quantity': row[5],
                    'price': row[6],
                    'timestamp': row[7]
                })
        
        return jsonify(products)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/delete_product/<string:product_id>', methods=['DELETE'])
def delete_product(product_id):
    try:
        files = init_data_files()
        wb = openpyxl.load_workbook(files['products'])
        ws = wb.active
        
        row_index = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1), 2):
            if row[0].value == product_id:
                row_index = idx
                break
        
        if row_index:
            ws.delete_rows(row_index)
            wb.save(files['products'])
            return jsonify(success=True)
        return jsonify(success=False, error="Product not found"), 404
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500

@app.route('/download_excel')
def download_excel():
    files = init_data_files()
    return send_file(files['products'], as_attachment=True)

# ... rest of the routes remain the same ...

if __name__ == '__main__':
    app.run(debug=True)