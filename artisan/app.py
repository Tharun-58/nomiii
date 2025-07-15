from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl import Workbook
import os
import datetime
import re
import random

app = Flask(__name__)

# Initialize Excel file
def init_excel():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    filepath = 'data/artisandata.xlsx'
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Product Name', 'Category', 'Description_EN', 'Description_TA', 'Quantity', 'Price', 'Timestamp'])
        wb.save(filepath)
    return filepath

# AI Description Generator
def generate_description(product_name, lang='EN'):
    descriptions = {
        'EN': [
            f"Beautiful handmade {product_name} crafted with traditional techniques",
            f"Authentic {product_name} featuring intricate designs and premium materials",
            f"Exquisite {product_name} showcasing local craftsmanship at its finest",
            f"Finely crafted {product_name} with attention to detail",
            f"Traditional {product_name} made with sustainable materials"
        ],
        'TA': [
            f"பாரம்பரியத் திறன்முறைகளால் உருவாக்கப்பட்ட அழகான கைவினை {product_name}",
            f"சிக்கலான வடிவமைப்புகள் மற்றும் உயர்தர பொருட்களுடன் உண்மையான {product_name}",
            f"உள்ளூர் கைவினைத் திறனை சிறப்பாக வெளிப்படுத்தும் அருமையான {product_name}",
            f"விவரங்களுக்கு கவனம் செலுத்தி நேர்த்தியாக உருவாக்கப்பட்ட {product_name}",
            f"நிலையான பொருட்களால் செய்யப்பட்ட பாரம்பரிய {product_name}"
        ]
    }
    return random.choice(descriptions[lang])

# Category Prediction Logic
def predict_category(product_name):
    textile_keywords = ['saree', 'pudavai', 'புடவை', 'silk', 'cotton', 'வேஷ்டி', 'veshti']
    clay_keywords = ['terracotta', 'pottery', 'clay', 'மண்', 'களிமண்', 'pot', 'பானை']
    metal_keywords = ['plate', 'bell', 'ஆணி', 'பித்தளை', 'வெண்கலம்', 'brass', 'statue', 'சிலை']
    jewelry_keywords = ['jewelry', 'நகை', 'bracelet', 'காப்பு', 'necklace', 'மாலை']
    
    product_lower = product_name.lower()
    
    if any(kw in product_lower for kw in textile_keywords):
        return 'Textiles'
    elif any(kw in product_lower for kw in clay_keywords):
        return 'Clay Art'
    elif any(kw in product_lower for kw in metal_keywords):
        return 'Metal Craft'
    elif any(kw in product_lower for kw in jewelry_keywords):
        return 'Jewelry'
    return 'Handicrafts'

# Extract price from voice input
def extract_price(text):
    # Tamil number words to digits
    tamil_numbers = {
        'ஒன்று': 1, 'இரண்டு': 2, 'மூன்று': 3, 'நான்கு': 4, 'ஐந்து': 5,
        'ஆறு': 6, 'ஏழு': 7, 'எட்டு': 8, 'ஒன்பது': 9, 'பத்து': 10,
        'இருபது': 20, 'முப்பது': 30, 'நாற்பது': 40, 'ஐம்பது': 50,
        'அறுபது': 60, 'எழுபது': 70, 'எண்பது': 80, 'தொண்ணூறு': 90,
        'நூறு': 100, 'இருநூறு': 200, 'முன்னூறு': 300, 'நாநூறு': 400,
        'ஐநூறு': 500, 'அறுநூறு': 600, 'எழுநூறு': 700, 'எண்ணூறு': 800,
        'தொள்ளாயிரம்': 900, 'ஆயிரம்': 1000
    }
    
    # Try to find numbers in text
    numbers = re.findall(r'\d+', text)
    if numbers:
        return int(numbers[-1])
    
    # Try Tamil number words
    for word, value in tamil_numbers.items():
        if word in text:
            return value
    
    # Fallback
    return 0

# Extract quantity from voice input
def extract_quantity(text):
    tamil_numbers = {
        'ஒன்று': 1, 'இரண்டு': 2, 'மூன்று': 3, 'நான்கு': 4, 'ஐந்து': 5,
        'ஆறு': 6, 'ஏழு': 7, 'எட்டு': 8, 'ஒன்பது': 9, 'பத்து': 10,
        'பதினொன்று': 11, 'பன்னிரண்டு': 12, 'பதிமூன்று': 13, 'பதினான்கு': 14,
        'பதினைந்து': 15, 'பதினாறு': 16, 'பதினேழு': 17, 'பதினெட்டு': 18,
        'பத்தொன்பது': 19, 'இருபது': 20
    }
    
    # Try to find numbers in text
    numbers = re.findall(r'\d+', text)
    if numbers:
        return int(numbers[0])
    
    # Try Tamil number words
    for word, value in tamil_numbers.items():
        if word in text:
            return value
    
    # Fallback
    return 1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add_product', methods=['POST'])
def add_product():
    filepath = init_excel()
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    data = request.json
    product_name = data['name']
    quantity = data['quantity']
    price = data['price']
    
    # Generate AI content
    category = predict_category(product_name)
    desc_en = generate_description(product_name, 'EN')
    desc_ta = generate_description(product_name, 'TA')
    
    # Add to Excel
    new_id = ws.max_row
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        new_id, 
        product_name, 
        category, 
        desc_en, 
        desc_ta, 
        quantity, 
        price, 
        timestamp
    ])
    
    wb.save(filepath)
    return jsonify(success=True, id=new_id)

@app.route('/get_products')
def get_products():
    filepath = init_excel()
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Skip empty rows
            products.append({
                'id': row[0],
                'name': row[1],
                'category': row[2],
                'desc_en': row[3],
                'desc_ta': row[4],
                'quantity': row[5],
                'price': row[6],
                'timestamp': row[7]
            })
    
    return jsonify(products)

@app.route('/delete_product/<int:product_id>', methods=['DELETE'])
def delete_product(product_id):
    filepath = init_excel()
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find the row to delete
    for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        if row[0].value == product_id:
            ws.delete_rows(idx)
            break
    
    wb.save(filepath)
    return jsonify(success=True)

@app.route('/download_excel')
def download_excel():
    init_excel()  # Ensure file exists
    return send_file('data/artisandata.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)