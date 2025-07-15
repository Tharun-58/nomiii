from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import random
import numpy as np
from collections import defaultdict

app = Flask(__name__)
CORS(app)

# File paths
EXCEL_FILE = 'deliverydata.xlsx'

# Create Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Deliveries"
        headers = [
            "DeliveryID", "Product", "Supplier", "Zone", 
            "ScheduledTime", "ActualTime", "OTP", "Status", "Delay(min)"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")
    else:
        print(f"Using existing Excel file: {EXCEL_FILE}")

# Check for delivery delay
def check_delay(scheduled, actual):
    delta = actual - scheduled
    delay_min = max(0, delta.total_seconds() / 60)
    status = "Delayed" if delay_min > 15 else "On Time"
    return status, delay_min

# Generate delivery ID
def generate_delivery_id(last_id):
    if last_id is None:
        return "NOMIID001"
    num = int(last_id[6:]) + 1
    return f"NOMIID{num:03d}"

# Get data from Excel
def get_excel_data():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Skip empty rows
            data.append({
                "DeliveryID": row[0],
                "Product": row[1],
                "Supplier": row[2],
                "Zone": row[3],
                "ScheduledTime": row[4],
                "ActualTime": row[5],
                "OTP": row[6],
                "Status": row[7],
                "Delay(min)": row[8]
            })
    return data

# Calculate metrics
def calculate_metrics(deliveries):
    if not deliveries:
        return {
            "total_deliveries": 0,
            "success_rate": 0,
            "on_time_rate": 0,
            "avg_delay": 0,
            "recent_deliveries": [],
            "delivery_trends": {"labels": [], "on_time": [], "delayed": []},
            "ai_alerts": []
        }
    
    # Basic metrics
    total = len(deliveries)
    successful = sum(1 for d in deliveries if d["Status"] != "Failed")
    on_time = sum(1 for d in deliveries if d["Status"] == "On Time")
    delays = [d["Delay(min)"] for d in deliveries if d["Status"] in ["Delayed", "On Time"]]
    
    # Delivery trends (last 7 days)
    today = datetime.now().date()
    trends = defaultdict(lambda: {"on_time": 0, "delayed": 0})
    
    # Initialize last 7 days
    for i in range(7):
        date_str = (today - timedelta(days=6-i)).strftime("%Y-%m-%d")
        trends[date_str] = {"on_time": 0, "delayed": 0}
    
    # Count deliveries per day
    for delivery in deliveries:
        if isinstance(delivery["ActualTime"], datetime):
            date_key = delivery["ActualTime"].strftime("%Y-%m-%d")
            if date_key in trends:
                if delivery["Status"] == "On Time":
                    trends[date_key]["on_time"] += 1
                elif delivery["Status"] == "Delayed":
                    trends[date_key]["delayed"] += 1
    
    # AI performance alerts
    alerts = []
    zone_delays = defaultdict(list)
    supplier_delays = defaultdict(list)
    
    # Collect delay data for zones and suppliers
    for delivery in deliveries:
        if delivery["Status"] in ["On Time", "Delayed"]:
            zone_delays[delivery["Zone"]].append(delivery["Delay(min)"])
            supplier_delays[delivery["Supplier"]].append(delivery["Delay(min)"])
    
    # Zone delay alerts
    for zone, delays in zone_delays.items():
        if len(delays) > 3 and np.mean(delays) > 20:
            alerts.append({
                "type": "warning",
                "title": "Repeated Delays",
                "message": f"Zone {zone} has average delay of {np.mean(delays):.1f} min over {len(delays)} deliveries"
            })
    
    # Supplier delay alerts
    for supplier, delays in supplier_delays.items():
        if len(delays) > 2 and np.mean(delays) > 25:
            alerts.append({
                "type": "warning",
                "title": "Supplier Delay Pattern",
                "message": f"Deliveries from {supplier} are delayed by {np.mean(delays):.1f} min on average"
            })
    
    # Low success rate alert
    success_rate = (successful / total) * 100
    if success_rate < 85:
        alerts.append({
            "type": "warning",
            "title": "Low Success Rate",
            "message": f"Your delivery success rate is {success_rate:.1f}%, below target of 85%"
        })
    
    # Positive alert when doing well
    if not alerts and success_rate > 95 and (on_time / total) > 0.9:
        alerts.append({
            "type": "info",
            "title": "Great Performance!",
            "message": "You're maintaining excellent delivery performance. Keep it up!"
        })
    
    # Traffic pattern alerts
    current_hour = datetime.now().hour
    if 16 <= current_hour <= 19:
        alerts.append({
            "type": "warning",
            "title": "Traffic Advisory",
            "message": "Heavy traffic expected during evening rush hours (4-7 PM)"
        })
    
    # Get recent deliveries (last 10)
    recent_deliveries = sorted(
        deliveries, 
        key=lambda x: x["ActualTime"] if isinstance(x["ActualTime"], datetime) else datetime.min,
        reverse=True
    )[:10]
    
    # Format recent deliveries for frontend
    formatted_recent = []
    for delivery in recent_deliveries:
        formatted = {
            "delivery_id": delivery["DeliveryID"],
            "product": delivery["Product"],
            "supplier": delivery["Supplier"],
            "zone": delivery["Zone"],
            "timestamp": delivery["ActualTime"].isoformat() if isinstance(delivery["ActualTime"], datetime) else "",
            "status": delivery["Status"],
            "delay": delivery["Delay(min)"]
        }
        formatted_recent.append(formatted)
    
    return {
        "total_deliveries": total,
        "success_rate": round((successful / total) * 100, 1) if total > 0 else 0,
        "on_time_rate": round((on_time / total) * 100, 1) if total > 0 else 0,
        "avg_delay": round(np.mean(delays), 1) if delays else 0,
        "recent_deliveries": formatted_recent,
        "delivery_trends": {
            "labels": list(trends.keys()),
            "on_time": [v["on_time"] for v in trends.values()],
            "delayed": [v["delayed"] for v in trends.values()]
        },
        "ai_alerts": alerts
    }

# Flask routes
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)

@app.route('/submit_delivery', methods=['POST'])
def submit_delivery():
    data = request.json
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    try:
        # Get last delivery ID
        last_id = ws.cell(row=ws.max_row, column=1).value if ws.max_row > 1 else None
        delivery_id = generate_delivery_id(last_id)
        
        # Calculate times
        scheduled_time = datetime.fromisoformat(data['scheduled_time'])
        actual_time = datetime.now()
        
        # Check OTP
        entered_otp = data.get('entered_otp', '')
        generated_otp = data.get('otp', '')
        
        if entered_otp != generated_otp:
            status = "Failed"
            delay_min = 0
        else:
            # Check delay
            status, delay_min = check_delay(scheduled_time, actual_time)
        
        # Create new entry
        new_row = [
            delivery_id,
            data['product'],
            data['supplier'],
            data['zone'],
            scheduled_time,
            actual_time,
            generated_otp,
            status,
            delay_min
        ]
        
        ws.append(new_row)
        wb.save(EXCEL_FILE)
        
        return jsonify({
            "status": "success",
            "delivery_id": delivery_id,
            "message": "Delivery recorded successfully"
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 400

@app.route('/metrics')
def get_metrics():
    try:
        deliveries = get_excel_data()
        metrics = calculate_metrics(deliveries)
        return jsonify(metrics)
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

if __name__ == '__main__':
    init_excel()
    print("Nomii Delivery Dashboard backend running...")
    print("Access the dashboard at: http://localhost:5000")
    app.run(debug=True, port=5000)